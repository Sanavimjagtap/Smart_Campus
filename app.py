import threading
from SmartCampus_AI.head_direction import get_head_direction
import time
import cv2
import os
from SmartCampus_AI.attendance_recognition import start_attendance
import re
from flask import Flask, render_template, request, redirect, url_for, session
from datetime import datetime
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font
from flask import send_file
import tempfile 
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from flask import jsonify

app = Flask(__name__)
app.secret_key = "smartcampus123"

classrooms = {

"AC-1":{
    "occupancy":"available",
    "temperature":0,
    "air":0,
    "lights":"OFF",
    "window":"Closed",
    "fire":"SAFE",
    "door":"Locked"
},

"AC-2":{
    "occupancy":"available",
    "temperature":0,
    "air":0,
    "lights":"OFF",
    "window":"Closed",
    "fire":"SAFE",
    "door":"Locked"
},

"AC-3":{
    "occupancy":"available",
    "temperature":0,
    "air":0,
    "lights":"OFF",
    "window":"Closed",
    "fire":"SAFE",
    "door":"Locked"
},
"AC-4":{
    "occupancy":"available",
    "temperature":0,
    "air":0,
    "lights":"OFF",
    "window":"Closed",
    "fire":"SAFE",
    "door":"Locked"
},
"AC-5":{
    "occupancy":"available",
    "temperature":0,
    "air":0,
    "lights":"OFF",
    "window":"Closed",
    "fire":"SAFE",
    "door":"Locked"
},
"AC-6":{
    "occupancy":"available",
    "temperature":0,
    "air":0,
    "lights":"OFF",
    "window":"Closed",
    "fire":"SAFE",
    "door":"Locked"
},
"AC-7":{
    "occupancy":"available",
    "temperature":0,
    "air":0,
    "lights":"OFF",
    "window":"Closed",
    "fire":"SAFE",
    "door":"Locked"
},
# Continue same for every room...

}


@app.route("/", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        login_id = request.form["login_id"]
        password = request.form["password"]

        connection = sqlite3.connect("smartcampus.db")
        cursor = connection.cursor()

        cursor.execute(
        "SELECT * FROM Teachers WHERE TeacherID=? AND Password=?",
        (login_id, password)
        )

        teacher = cursor.fetchone()



        cursor.execute(
        "SELECT * FROM Students WHERE StudentID=? AND Password=?",
        (login_id, password)
        )

        student = cursor.fetchone()
        connection.close()
        if teacher:
            session["teacher_id"] = teacher[0]
            session["teacher_name"] = teacher[1]
            session["department"] = teacher[3]
            return redirect(url_for("teacher"))
        
        elif student:

            session["student_id"] = student[0]
            session["student_name"] = student[1]
            session["class"] = student[3]
            session["division"] = student[4]

            connection.close()
            return redirect(url_for("student"))
        
        else:
            return "Invalid Login ID"

    return render_template("Login.html")



@app.route("/teacher")
def teacher():

    return render_template(
        "teacher1.html",
        teacher_name=session["teacher_name"],
        active_page="dashboard"
    )
   


@app.route("/student")
def student():

    return render_template(
        "Student_page.html",
        student_name=session["student_name"]
    )

@app.route("/session_setup")
def session_setup():
	return render_template(
        "session_setup.html",
		active_page="session"
	)

@app.route("/session", methods=["POST"])
def session_page():
    
    standard = request.form["std"]
    division = request.form["div"]
    duration = request.form["time"]
    room = request.form["room"]
    now = datetime.now()
    current_date = now.strftime("%d %B %Y")
    current_time = now.strftime("%I:%M:%S %p")

    minutes=int(request.form["time"])
    seconds=minutes*60  #10

    connection = sqlite3.connect("smartcampus.db")
    cursor = connection.cursor()

    cursor.execute("""
    INSERT INTO Sessions
    (
    TeacherID,
    Class,
    Division,
    Room,
    Date,
    StartTime,
    EndTime,
    Duration
    )

    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """,
    (
    session["teacher_id"],
    standard,
    division,
    room,
    current_date,
    current_time,
    "",
    minutes
    )
    )
    
    session_id = cursor.lastrowid
    session["session_id"] = session_id
    connection.commit()
    connection.close()
    threading.Thread(
        target=start_attendance,
        args=(session_id,),
        daemon=True
    ).start()
    
    return render_template(
        "session.html",
        active_page="session",
    standard=standard,
    division=division,
    duration=duration,
    time=current_time,
    date=current_date,
    seconds=seconds,
    session_id=session_id,
    room=room,
)

@app.route("/present_count/<int:session_id>")
def present_count(session_id):

    connection = sqlite3.connect("smartcampus.db")
    cursor = connection.cursor()

    cursor.execute("""
    SELECT COUNT(*)
    FROM Attendance
    WHERE SessionID = ?
    """, (session_id,))

    count = cursor.fetchone()[0]

    connection.close()

    return {
        "count": count
    }

@app.route("/end_session", methods=["POST"])
def end_session():
    session_id = request.form["session_id"]
    now = datetime.now()

    current_time = now.strftime("%I:%M:%S %p")
    connection = sqlite3.connect("smartcampus.db")

    cursor = connection.cursor()
    cursor.execute("""
    UPDATE Sessions
    SET EndTime = ?
    WHERE SessionID = ?
    """,
    (current_time, session_id)
    )
    connection.commit()
    connection.close()
    return redirect(url_for("teacher"))

@app.route("/checkin", methods=["POST"])
def checkin():

    now = datetime.now()

    current_date = now.strftime("%d %B %Y")
    current_time = now.strftime("%I:%M:%S %p")

    connection = sqlite3.connect("smartcampus.db")
    cursor = connection.cursor()

    # 👇 STEP 1: Check if teacher has already checked in today
    cursor.execute("""
    SELECT * FROM TeacherAttendance
    WHERE TeacherID=? AND Date=?
    """,
    (
        session["teacher_id"],
        current_date
    ))

    attendance = cursor.fetchone()

    # 👇 STEP 2: If record already exists, don't insert again
    if attendance:
        connection.close()
        return "Already checked in today."

    # 👇 STEP 3: Otherwise insert a new record
    cursor.execute("""
    INSERT INTO TeacherAttendance
    (TeacherID, Date, CheckIn)
    VALUES (?, ?, ?)
    """,
    (
        session["teacher_id"],
        current_date,
        current_time
    ))

    connection.commit()
    connection.close()

    return redirect(url_for("teacher"))


@app.route("/checkout", methods=["POST"])
def checkout():

    now = datetime.now()

    current_date = now.strftime("%d %B %Y")
    current_time = now.strftime("%I:%M:%S %p")

    connection = sqlite3.connect("smartcampus.db")
    cursor = connection.cursor()
    cursor.execute("""
    UPDATE TeacherAttendance
    SET CheckOut = ?
    WHERE TeacherID = ? AND Date = ?
    """,
    (
        current_time,
        session["teacher_id"],
        current_date
    ))
    connection.commit()
    connection.close()

    return redirect(url_for("teacher"))

@app.route("/mark_attendance/<student_id>")
def mark_attendance(student_id):

    now = datetime.now()
    current_time = now.strftime("%I:%M:%S %p")

    connection = sqlite3.connect(
        os.path.join(BASE_DIR, "smartcampus.db"), 
        check_same_thread=False
    )
    cursor = connection.cursor()

    cursor.execute("""
    SELECT * FROM Attendance
    WHERE SessionID=? AND StudentID=?
    """,
    (
    session["session_id"],
    student_id
    ))

    record = cursor.fetchone()

    if record:
        connection.close()
        return "Already Marked"
    
    cursor.execute("""
    INSERT INTO Attendance
    (SessionID, StudentID, TimeMarked)
    VALUES (?, ?, ?)
    """,
    (
        session["session_id"],
        student_id,
        current_time
    ))

    connection.commit()
    connection.close()

    return "Attendance Marked!"

@app.route("/register_face")
def register_face():

    connection = sqlite3.connect("smartcampus.db")
    cursor = connection.cursor()

    cursor.execute("""
    SELECT StudentID, Name
    FROM Students
    """)

    students = cursor.fetchall()

    connection.close()

    return render_template(
        "register_face.html",
        students=students,
        active_page="face"
    )
@app.route("/capture_face", methods=["POST"])
def capture_face():

    student_id = request.form["student_id"]

    folder = f"dataset/{student_id}"

    if not os.path.exists(folder):
        os.makedirs(folder)

    camera = cv2.VideoCapture(1, cv2.CAP_MSMF)

    cv2.namedWindow("Student Face Registration", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("Student Face Registration", 900, 700)

    face_detector = cv2.CascadeClassifier(
        cv2.data.haarcascades +
        "haarcascade_frontalface_default.xml"
    )

    count = 0
    last_capture = time.time()

    while True:

        ret, frame = camera.read()

        if not ret:
            break

        direction = get_head_direction(frame)

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        faces = face_detector.detectMultiScale(
            gray,
            1.2,
            5
        )

        # ---------- Decide required pose ----------

        if count < 20:
            required_pose = "Straight"
            instruction = "Look Straight"

        elif count < 40:
            required_pose = "Left"
            instruction = "Turn Left"

        elif count < 60:
            required_pose = "Right"
            instruction = "Turn Right"

        elif count < 80:
            required_pose = "Straight"
            instruction = "Look Up (Coming Soon)"

        else:
            required_pose = "Straight"
            instruction = "Smile :)"

        # ---------- Face ----------

        for (x, y, w, h) in faces:

            face = gray[y:y+h, x:x+w]

            current_time = time.time()

            if direction == required_pose:

                if current_time - last_capture > 0.2:

                    count += 1

                    cv2.imwrite(
                        f"{folder}/{count}.jpg",
                        face
                    )

                    last_capture = current_time

            cv2.rectangle(
                frame,
                (x, y),
                (x+w, y+h),
                (0,255,0),
                2
            )

        # ---------- Display ----------

        cv2.putText(
            frame,
            f"Images : {count}/100",
            (20,40),
            cv2.FONT_HERSHEY_SIMPLEX,
            1,
            (0,255,0),
            2
        )

        cv2.putText(
            frame,
            instruction,
            (20,80),
            cv2.FONT_HERSHEY_SIMPLEX,
            1,
            (255,0,0),
            2
        )

        color = (0,255,0)

        if direction != required_pose:
            color = (0,0,255)

        cv2.putText(
            frame,
            f"Detected : {direction}",
            (20,120),
            cv2.FONT_HERSHEY_SIMPLEX,
            1,
            color,
            2
        )

        if direction == required_pose:

            cv2.putText(
                frame,
                "CAPTURING...",
                (20,170),
                cv2.FONT_HERSHEY_SIMPLEX,
                1,
                (0,255,0),
                2
            )

        else:

            cv2.putText(
                frame,
                "Waiting...",
                (20,170),
                cv2.FONT_HERSHEY_SIMPLEX,
                1,
                (0,0,255),
                2
            )

        cv2.setWindowProperty(
            "Student Face Registration",
            cv2.WND_PROP_TOPMOST,
            1
        )

        cv2.imshow(
            "Student Face Registration",
            frame
        )

        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

        if count >= 100:
            break

    camera.release()
    cv2.destroyAllWindows()

    return redirect(url_for("teacher"))

@app.route("/attendance")
def attendance():

    connection = sqlite3.connect("smartcampus.db")
    cursor = connection.cursor()

    # Latest session
    cursor.execute("""
    SELECT SessionID, Class, Division, Date
    FROM Sessions
    ORDER BY SessionID DESC
    LIMIT 1
    """)

    session_data = cursor.fetchone()

    if session_data is None:
        connection.close()
        return "No session found."

    session_id = session_data[0]

    # Get every student with attendance
    cursor.execute("""
    SELECT
        Students.StudentID,
        Students.Name,
        Attendance.TimeMarked
    FROM Students

    LEFT JOIN Attendance
    ON Students.StudentID = Attendance.StudentID
    AND Attendance.SessionID = ?

    ORDER BY Students.StudentID
    """,(session_id,))

    students = cursor.fetchall()

    connection.close()

    return render_template(
        "attendance.html",
        session_data=session_data,
        students=students,
        active_page="attendance"
    )

@app.route("/attendance_report/<int:session_id>")
def attendance_report(session_id):

    connection = sqlite3.connect("smartcampus.db")
    cursor = connection.cursor()

    cursor.execute("""
    SELECT Class, Division, Date, StartTime
    FROM Sessions
    WHERE SessionID=?
    """,(session_id,))

    session_info = cursor.fetchone()

    cursor.execute("""
    SELECT
        Students.StudentID,
        Students.Name,
        Attendance.TimeMarked
    FROM Students

    LEFT JOIN Attendance
    ON Students.StudentID = Attendance.StudentID
    AND Attendance.SessionID=?

    ORDER BY Students.StudentID
    """,(session_id,))

    students = cursor.fetchall()

    present = sum(1 for s in students if s[2] is not None)
    total = len(students)
    absent = total - present

    connection.close()

    return render_template(
        "attendance_report.html",
        students=students,
        session_id=session_id,
        session_info=session_info,
        present=present,
        absent=absent,
        total=total
    )

@app.route("/attendance_reports")
def attendance_reports():

    connection = sqlite3.connect("smartcampus.db")
    cursor = connection.cursor()

    cursor.execute("""
    SELECT
        SessionID,
        Class,
        Division,
        Date,
        StartTime
    FROM Sessions
    ORDER BY SessionID DESC
    """)

    sessions = cursor.fetchall()

    connection.close()

    return render_template(
        "attendance_reports.html",
        sessions=sessions
    )
@app.route("/export_excel/<int:session_id>")
def export_excel(session_id):

    connection = sqlite3.connect("smartcampus.db")
    cursor = connection.cursor()

    # Session details
    cursor.execute("""
    SELECT Class, Division, Date
    FROM Sessions
    WHERE SessionID=?
    """, (session_id,))

    session_data = cursor.fetchone()

    # Attendance list
    cursor.execute("""
    SELECT
        Students.StudentID,
        Students.Name,
        Attendance.TimeMarked

    FROM Students

    LEFT JOIN Attendance
    ON Students.StudentID = Attendance.StudentID
    AND Attendance.SessionID = ?

    ORDER BY Students.StudentID
    """, (session_id,))

    students = cursor.fetchall()

    connection.close()

    wb = Workbook()

    ws = wb.active

    ws.title = "Attendance"

    # Heading
    ws["A1"] = "FERGUSSON COLLEGE"
    ws["A1"].font = Font(size=18, bold=True)

    ws["A3"] = f"Class : {session_data[0]}"
    ws["B3"] = f"Division : {session_data[1]}"
    ws["C3"] = f"Date : {session_data[2]}"

    # Table headings
    ws.append([])

    ws.append([
        "Student ID",
        "Student Name",
        "Status",
        "Time Marked"
    ])

    for cell in ws[5]:
        cell.font = Font(bold=True)

    # Student data
    for student in students:

        status = "Present" if student[2] else "Absent"

        ws.append([
            student[0],
            student[1],
            status,
            student[2] if student[2] else "-"
        ])

    # Adjust column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20

    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")

    wb.save(temp.name)

    return send_file(
        temp.name,
        as_attachment=True,
        download_name=f"Attendance_Session_{session_id}.xlsx"
    )

@app.route("/export_pdf/<int:session_id>")
def export_pdf(session_id):

    connection = sqlite3.connect("smartcampus.db")
    cursor = connection.cursor()

    cursor.execute("""
    SELECT Class, Division, Date
    FROM Sessions
    WHERE SessionID=?
    """,(session_id,))

    session_data = cursor.fetchone()

    cursor.execute("""
    SELECT
        Students.StudentID,
        Students.Name,
        Attendance.TimeMarked

    FROM Students

    LEFT JOIN Attendance
    ON Students.StudentID=Attendance.StudentID
    AND Attendance.SessionID=?

    ORDER BY Students.StudentID
    """,(session_id,))

    students = cursor.fetchall()

    connection.close()

    temp = tempfile.NamedTemporaryFile(delete=False,suffix=".pdf")

    doc = SimpleDocTemplate(temp.name)

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph("<b>FERGUSSON COLLEGE</b>",styles["Title"])
    )

    elements.append(
        Paragraph("<b>Smart Campus Attendance Report</b>",styles["Heading2"])
    )

    elements.append(
        Paragraph(
            f"Class : {session_data[0]}<br/>"
            f"Division : {session_data[1]}<br/>"
            f"Date : {session_data[2]}",
            styles["BodyText"]
        )
    )

    elements.append(Paragraph("<br/>",styles["BodyText"]))

    data = [["Student ID","Name","Status","Time"]]

    for student in students:

        status = "Present" if student[2] else "Absent"

        data.append([
            student[0],
            student[1],
            status,
            student[2] if student[2] else "-"
        ])

    table = Table(data,colWidths=[1.3*inch,2.3*inch,1.1*inch,1.4*inch])

    table.setStyle(TableStyle([

        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#75619D")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),

        ("GRID",(0,0),(-1,-1),1,colors.grey),

        ("BACKGROUND",(0,1),(-1,-1),colors.whitesmoke),

        ("ALIGN",(0,0),(-1,-1),"CENTER"),

        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ("BOTTOMPADDING",(0,0),(-1,0),10),

    ]))

    elements.append(table)

    doc.build(elements)

    return send_file(
        temp.name,
        as_attachment=True,
        download_name=f"Attendance_Session_{session_id}.pdf"
    )

@app.route("/occupancy")
def occupancy_page():
    print("PAGE PID:", os.getpid())
    print(classrooms)
    room_list = []

    for room_name, data in classrooms.items():

        room = data.copy()

        room["room"] = room_name

        room_list.append(room)

    return render_template(
        "occupancy.html",
        classrooms=room_list,
        active_page="occupancy"
    )

@app.route("/chatbot")
def chatbot():

    return render_template(
        "chatbot.html",
        active_page="chatbot"
    )

@app.route("/ask_chatbot", methods=["POST"])
def ask_chatbot():

    question = request.form["question"].lower()

    # --------------------
    # Library
    # --------------------
    if "library" in question:

        answer = """
The Fergusson College Library is open from
8:00 AM to 8:00 PM on weekdays.

You can issue books using your student ID.
"""

    # --------------------
    # Faculty
    # --------------------
    elif "faculty" in question:

        answer = """
Departments available:

• Physics
• Chemistry
• Mathematics
• Computer Science
• Electronics
"""

    # --------------------
    # Labs
    # --------------------
    elif "lab" in question:

        answer = """
Laboratories Available:

• Computer Lab 1
• Computer Lab 2
• Electronics Lab 1
• Electronics Lab 2
"""

    # --------------------
    # Admissions
    # --------------------
    elif "admission" in question:

        answer = """
Admissions are conducted through the Fergusson College admission portal.

For detailed eligibility, visit the college website.
"""

    # --------------------
    # AC Rooms
    # --------------------
    elif re.search(r'ac[\s-]?\d+', question) or "classroom" in question:

        room = re.search(r'ac[\s-]?\d+', question)

        if room:

            room = room.group().upper().replace(" ", "-")

            if room in ["AC1","AC-1"]:
                floor = "First Floor"

            elif room in ["AC2","AC-2"]:
                floor = "First Floor"

            elif room in ["AC3","AC-3"]:
                floor = "First Floor"

            elif room in ["AC4","AC-4"]:
                floor = "First Floor"

            elif room in ["AC5","AC-5"]:
                floor = "First Floor"

            elif room in ["AC6","AC-6"]:
                floor = "First Floor"

            elif room in ["AC7","AC-7"]:
                floor = "Second Floor"

            elif room in ["AC8","AC-8"]:
                floor = "Second Floor"

            elif room in ["AC9","AC-9"]:
                floor = "Second Floor"

            elif room in ["AC10","AC-10"]:
                floor = "Second Floor"

            elif room in ["AC11","AC-11"]:
                floor = "Second Floor"

            elif room in ["AC12","AC-12"]:
                floor = "Second Floor"

            else:
                floor = "Unknown"

            answer = f"{room} is located on the {floor}."

        else:

            answer = """
    AC-1 to AC-6 → First Floor

    AC-7 to AC-12 → Second Floor
    """

    else:

        answer = """
Sorry, I couldn't understand that.

Please ask about:

• Library
• Faculty
• Labs
• Admissions
• Campus Map
"""

    return jsonify({
        "answer": answer
    })

@app.route("/update_classroom", methods=["POST"])
def update_classroom():

    data=request.json

    print("Received:", data)

    if not data or "room" not in data:
        return "Invalid Data",400


    room=data["room"]

    classrooms[room]=data

    return "Updated"

@app.route("/occupancy_status")
def occupancy_status():

    return jsonify(classrooms)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)

    
